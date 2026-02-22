import os
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from shop.models import User, PickupPoint, Product, Order, OrderItem

class Command(BaseCommand):
    help = 'Импортирует данные из Excel (Прил_2)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Путь к файлу Excel')

    @transaction.atomic
    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        wb = openpyxl.load_workbook(file_path, data_only=True)

        self.stdout.write('Очистка базы данных...')
        OrderItem.objects.all().delete()
        Order.objects.all().delete()
        Product.objects.all().delete()
        User.objects.filter(is_superuser=False).delete()
        PickupPoint.objects.all().delete()

        # Импорт Pickup Points
        self.stdout.write('Импорт Пунктов выдачи...')
        ws_pp = wb['Pickup point']
        pp_dict = {}
        for idx, row in enumerate(ws_pp.iter_rows(min_row=2, values_only=True), start=1):
            if not row[0]:
                continue
            pp = PickupPoint.objects.create(address=row[0])
            pp_dict[idx] = pp

        # Импорт Users
        self.stdout.write('Импорт Пользователей...')
        ws_user = wb['User']
        user_dict = {}
        role_map = {
            'Администратор': 'admin',
            'Менеджер': 'manager',
            'Клиент': 'client',
            'Гость': 'guest',
        }
        for row in ws_user.iter_rows(min_row=2, values_only=True):
            role_ru, fio, login, password = row
            if not login:
                continue
            parts = fio.split()
            last_name = parts[0] if len(parts) > 0 else ''
            first_name = parts[1] if len(parts) > 1 else ''
            patronymic = parts[2] if len(parts) > 2 else ''

            user = User.objects.create_user(
                username=login,
                password=password,
                first_name=first_name,
                last_name=last_name,
                role=role_map.get(role_ru, 'guest')
            )
            user.patronymic = patronymic
            if user.role == 'admin':
                user.is_staff = True
                user.is_superuser = True
            user.save()
            user_dict[fio] = user

        # Импорт Products
        self.stdout.write('Импорт Товаров...')
        ws_prod = wb['Product']
        product_dict = {}
        for row in ws_prod.iter_rows(min_row=2, values_only=True):
            sku = row[0]
            if not sku:
                continue
            product = Product.objects.create(
                sku=str(sku),
                name=str(row[1] or ''),
                unit=str(row[2] or ''),
                price=row[3] or 0.0,
                supplier=str(row[4] or ''),
                manufacturer=str(row[5] or ''),
                category=str(row[6] or ''),
                discount=row[7] or 0,
                stock=row[8] or 0,
                description=str(row[9] or '')
            )
            product_dict[sku] = product
            # photo we ignore for now, we'll give a default stub if not exist
        
        # Импорт Orders
        self.stdout.write('Импорт Заказов...')
        ws_order = wb['Order']
        for row in ws_order.iter_rows(min_row=2, values_only=True):
            order_id = row[0]
            if not order_id:
                break
            
            items_str = str(row[1] or '')
            date_order = row[2]
            date_delivery = row[3]
            pp_idx = row[4]
            client_fio = row[5]
            code = row[6]
            status = row[7]

            pp = pp_dict.get(pp_idx)
            client = user_dict.get(client_fio)

            order = Order.objects.create(
                order_id=order_id,
                order_date=date_order,
                delivery_date=date_delivery,
                pickup_point=pp,
                client=client,
                code=code,
                status=status
            )

            if items_str:
                parts = [p.strip() for p in items_str.split(',')]
                for i in range(0, len(parts), 2):
                    if i + 1 < len(parts):
                        item_sku = parts[i]
                        quantity = int(parts[i+1])
                        prod = product_dict.get(item_sku)
                        if prod:
                            OrderItem.objects.create(
                                order=order,
                                product=prod,
                                quantity=quantity
                            )

        self.stdout.write(self.style.SUCCESS('Данные успешно импортированы!'))
